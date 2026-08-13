#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""回归测试：4 个文件处理工具的批量参数（v1.0.21 批量参数 + v1.0.22 框架并行下发提速）。

把此前在容器内临时验证的逻辑固化为可入库测试，确保批量参数行为不被回归：
  * read_file     -> paths 列表（多文件带分隔头拼接返回）
  * terminal_run  -> commands 列表（合并为单次执行，共享一次快照/进程）
  * code_exec     -> snippets 列表（单子进程内顺序执行，变量跨段保留，消除多次冷启动）
  * generate_excel-> files 列表（多工作簿，每个返回 /media 下载链接）

直接运行：  python tests/test_batch_tools.py
也可被 pytest 收集（同步 test_* 函数，无需 pytest-asyncio）。
"""
from __future__ import annotations

import asyncio
import json
import os
import shutil
import sys
import tempfile
from types import SimpleNamespace

# ── 让脚本在容器内（/app/backend 在 PYTHONPATH）或仓库根运行都能 import zhishu ──
_HERE = os.path.dirname(os.path.abspath(__file__))
for _cand in ("/app/backend", os.path.join(_HERE, "..", "backend")):
    _cand = os.path.abspath(_cand)
    if os.path.isdir(os.path.join(_cand, "zhishu")) and _cand not in sys.path:
        sys.path.insert(0, _cand)


# --------------------------------------------------------------------------
# 进程内 fake 装置（不触碰真实配置 / 媒体库）
# --------------------------------------------------------------------------
class FakeSecurityCode:
    allow_code_exec = True
    code_exec_timeout = 30
    code_exec_mem_limit_mb = 0
    code_exec_network_isolated = False


class FakeSecurityShell:
    allow_shell = True
    shell_allowlist = None
    shell_enforce_allowlist = False  # 仅走高危拦截，放行白名单外的只读命令


class FakeSecurityShellEnforce:
    allow_shell = True
    shell_allowlist = None
    shell_enforce_allowlist = True   # 同时走白名单 + 高危拦截


class FakeMedia:
    """极简媒体库替身：落盘到临时目录并返回 /media/... URL，记录已保存文件。"""
    def __init__(self, root):
        self.root = root
        os.makedirs(root, exist_ok=True)
        self.saved = {}  # basename -> 绝对路径

    def save_file(self, src_path, kind=None, owner=None):
        owner = owner or "tester"
        odir = os.path.join(self.root, owner)
        os.makedirs(odir, exist_ok=True)
        dst = os.path.join(odir, os.path.basename(src_path))
        shutil.copyfile(src_path, dst)
        self.saved[os.path.basename(dst)] = dst
        return f"/media/{owner}/{os.path.basename(dst)}"


def _make_ctx(security=None, media=None, user="tester", role="user"):
    from zhishu.core.tools.base import ToolContext
    return ToolContext(user=user, user_role=role, is_admin=False,
                       security=security, media=media)


def _install_fake_ctx(data_dir, store_dir="media"):
    """替换全局 get_ctx，让 read_file 的 _resolve_read_path 拿到可控的 media_root。"""
    import zhishu.context as ctxmod
    fake = SimpleNamespace(cfg=SimpleNamespace(
        server=SimpleNamespace(data_dir=data_dir),
        media=SimpleNamespace(store_dir=store_dir),
    ))
    ctxmod.get_ctx = lambda: fake
    return fake


# --------------------------------------------------------------------------
# 各工具检查
# --------------------------------------------------------------------------
async def check_read_file(owner="tester"):
    from zhishu.core.tools.builtins.file import read_file

    tmp = tempfile.mkdtemp(prefix="zh_read_")
    try:
        data_dir = os.path.join(tmp, "data")
        media_root = os.path.join(data_dir, "media")
        os.makedirs(os.path.join(media_root, owner), exist_ok=True)
        _install_fake_ctx(data_dir, "media")

        fa = os.path.join(media_root, owner, "a.txt")
        with open(fa, "w", encoding="utf-8") as f:
            f.write("alpha line1\nalpha line2")
        fb = os.path.join(media_root, owner, "b.txt")
        with open(fb, "w", encoding="utf-8") as f:
            f.write("beta only")

        ctx = _make_ctx()
        # 单文件向后兼容：不带分隔头
        single = await read_file({"path": f"/media/{owner}/a.txt"}, ctx)
        assert "alpha line1" in single, single
        assert "===== 文件" not in single, single
        # v1.0.24：单数调用返回须带【提速提示】引导改用 paths 批量
        assert "[提速提示]" in single, single

        # 批量 paths：带分隔头拼接
        multi = await read_file(
            {"paths": [f"/media/{owner}/a.txt", f"/media/{owner}/b.txt"]}, ctx)
        assert "===== 文件 a.txt =====" in multi, multi
        assert "===== 文件 b.txt =====" in multi, multi
        assert "alpha line1" in multi and "beta only" in multi, multi

        # 缺失文件：友好报错而非崩溃
        miss = await read_file({"paths": [f"/media/{owner}/nope.txt"]}, ctx)
        assert "文件不存在或越权" in miss, miss
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


async def check_terminal():
    from zhishu.core.tools.builtins.terminal import terminal_run

    # 批量 commands：合并为单次执行
    ctx = _make_ctx(security=FakeSecurityShell(), media=None)
    out = await terminal_run({"commands": ["echo hello", "echo world"]}, ctx)
    assert "hello" in out and "world" in out, out

    # 单命令向后兼容
    one = await terminal_run({"command": "echo solo"}, ctx)
    assert "solo" in one, one

    # 批量路径下安全门依然生效：高危命令被拦截
    blocked = await terminal_run(
        {"commands": ["rm -rf /"], "timeout": 10},
        _make_ctx(security=FakeSecurityShellEnforce(), media=None),
    )
    assert "[已拦截]" in blocked, blocked


async def check_code_exec():
    from zhishu.core.tools.builtins.code_exec import code_exec

    ctx = _make_ctx(security=FakeSecurityCode(), media=None)
    # snippets：单子进程顺序执行，变量跨段保留
    out = await code_exec({"snippets": ["x = 21 * 2", "print('RESULT', x)"]}, ctx)
    assert "RESULT 42" in out, out

    # 单 code 向后兼容
    one = await code_exec({"code": "print('SOLO_CODE')"}, ctx)
    assert "SOLO_CODE" in one, one
    # v1.0.24：单数调用返回须带【提速提示】引导改用 snippets 批量
    assert "[提速提示]" in one, one


async def check_generate_excel():
    from zhishu.core.tools.builtins.generate_excel import generate_excel
    import openpyxl

    tmp = tempfile.mkdtemp(prefix="zh_xlsx_")
    try:
        media = FakeMedia(os.path.join(tmp, "media"))
        ctx = _make_ctx(media=media)

        # 批量 files：多工作簿，每个返回 /media 链接
        out = await generate_excel({"files": [
            {"filename": "a.xlsx", "sheets": [{"name": "S", "rows": [["x", "y"], [1, 2]]}]},
            {"filename": "b.xlsx", "sheets": [{"name": "T", "rows": [["p"], [3]]}]},
        ]}, ctx)
        assert out.count("已生成标准 Excel 文件") == 2, out
        assert "/media/tester/" in out, out

        # 校验真实生成的文件可被 openpyxl 打开（确为合法 xlsx，非损坏）
        assert len(media.saved) == 2, media.saved
        for p in media.saved.values():
            wb = openpyxl.load_workbook(p)
            assert wb.sheetnames, p

        # 单工作簿向后兼容
        one = await generate_excel(
            {"filename": "c.xlsx", "sheets": [{"name": "U", "rows": [["q"]]}]}, ctx)
        assert "已生成标准 Excel 文件「c.xlsx」" in one, one
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


async def check_generate_excel_from_file():
    """v1.0.28：generate_excel 的 from_file 桥接 + 沙箱最近表格兜底 + 可操作报错。"""
    from zhishu.core.tools.builtins.generate_excel import (
        generate_excel, _resolve_from_file, _load_sheet_spec_from_file,
        _find_recent_sandbox_table,
    )
    import openpyxl

    tmp = tempfile.mkdtemp(prefix="zh_xf_")
    try:
        media = FakeMedia(os.path.join(tmp, "media"))
        ctx = _make_ctx(media=media)

        # 1) from_file 传 /media 链接（模拟 code_exec 自动发布的 CSV）
        csv_path = os.path.join(tmp, "years.csv")
        with open(csv_path, "w", encoding="utf-8") as f:
            f.write("年份,事件\n1914年,美国股市曾关门停市\n1919年,可口可乐上市\n")
        url = media.save_file(csv_path, kind="file", owner="tester")
        out = await generate_excel({"filename": "years.xlsx", "from_file": url}, ctx)
        assert "已生成标准 Excel 文件「years.xlsx」" in out, out
        xlsx = [p for p in media.saved if p.endswith(".xlsx")][0]
        wb = openpyxl.load_workbook(media.saved[xlsx])
        vals = [r[0].value for r in wb.active.iter_rows()]
        assert "1914年" in vals, vals

        # 2) 单元：_find_recent_sandbox_table 找到刚写入的表格文件
        from zhishu.core.tools.builtins.sandbox import sandbox_cwd_for
        sb = sandbox_cwd_for("tester")
        probe = os.path.join(sb, "probe.csv")
        with open(probe, "w", encoding="utf-8") as f:
            f.write("a,b\n1,2\n")
        try:
            found = _find_recent_sandbox_table("tester")
            assert found and os.path.basename(found) == "probe.csv", found
            # 沙箱兜底 end-to-end：只传 filename，应自动采用该 csv
            out2 = await generate_excel({"filename": "auto.xlsx"}, ctx)
            assert "已生成标准 Excel 文件「auto.xlsx」" in out2, out2
            assert "自动采用沙箱最近生成的表格文件" in out2, out2
        finally:
            if os.path.exists(probe):
                os.remove(probe)

        # 3) 既无数据也无文件：可操作报错（须含 from_file 指引，而非只报缺数据）
        out3 = await generate_excel({"filename": "empty.xlsx"}, ctx)
        assert "缺少表格数据" in out3 and "from_file" in out3, out3

        # 4) 单元：_resolve_from_file / _load_sheet_spec_from_file 基础行为
        assert _resolve_from_file(os.path.abspath(csv_path), media, "tester") is not None
        assert _resolve_from_file("/no/such/file.csv", media, "tester") is None
        spec = _load_sheet_spec_from_file(csv_path)
        assert spec and spec[0]["csv"].startswith("年份,事件"), spec
        # JSON list[dict] -> 自动表头
        jp = os.path.join(tmp, "rows.json")
        with open(jp, "w", encoding="utf-8") as f:
            json.dump([{"年份": "2000年", "事件": "x"}, {"年份": "2001年", "事件": "y"}], f)
        spec2 = _load_sheet_spec_from_file(jp)
        assert spec2 and spec2[0]["header"] == ["年份", "事件"], spec2
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


async def check_speed_hints():
    """v1.0.24：4 个文件处理工具的顶层 description 须以【提速关键】开头，
    把「用批量参数压 N 次同类调用为 1 次」这条最可靠的提速手段直接钉在模型决策点。"""
    from zhishu.core.tools.registry import ToolRegistry

    for name in ("read_file", "terminal_run", "code_exec", "generate_excel"):
        spec = ToolRegistry.get(name)
        assert spec is not None, f"{name} 未注册"
        assert spec.description.startswith("【提速关键】"), (
            f"{name} 的 description 未前置【提速关键】：{spec.description[:40]}")


async def check_schema_list_and_scalar():
    """v1.0.24 契约：4 个文件处理工具的 JSON Schema 同时暴露「标量主用 + 列表可选」两组参数
    （read_file: path+paths / terminal_run: command+commands / code_exec: code+snippets /
    generate_excel: filename+sheets+files）。标量参数保证模型以最可靠方式单调用（不诱发非法 JSON），
    列表参数保留批量能力。两种形态都必须存在，缺一不可。"""
    from zhishu.core.tools.registry import ToolRegistry

    checks = {
        "read_file": (("path", "paths"),),
        "terminal_run": (("command", "commands"),),
        "code_exec": (("code", "snippets"), "snippets"),
        "generate_excel": (("filename", "files"),),
    }
    for name, args in checks.items():
        spec = ToolRegistry.get(name)
        props = spec.parameters.get("properties", {})
        for scalar, listp in args[0:1]:
            assert scalar in props, f"{name} 缺少标量参数 '{scalar}'"
            assert listp in props, f"{name} 缺少列表参数 '{listp}'"
    # code_exec 必须同时有 code（标量主用）与 snippets（可选批量）
    ce = ToolRegistry.get("code_exec").parameters["properties"]
    assert "code" in ce and "snippets" in ce, "code_exec 须同时暴露 code 与 snippets"


async def check_consolidation_nudge():
    """v1.0.27：同工具多次调用「合并提醒」阈值与文案正确（不改 schema，零回归）。"""
    from zhishu.core.agent.agent import _build_consolidation_nudge

    # 未达阈值：返回 None，不注入
    assert _build_consolidation_nudge("code_exec", 1) is None
    assert _build_consolidation_nudge("code_exec", 2) is None
    assert _build_consolidation_nudge("read_file", 2) is None
    assert _build_consolidation_nudge("generate_excel", 1) is None
    # 达阈值：返回一次性系统提醒，且含工具名与合并指引
    n = _build_consolidation_nudge("code_exec", 3)
    assert n is not None and "code_exec" in n and "合并" in n
    n2 = _build_consolidation_nudge("generate_excel", 2)
    assert n2 is not None and "generate_excel" in n2 and "files" in n2
    n3 = _build_consolidation_nudge("read_file", 3)
    assert n3 is not None and "paths" in n3
    n4 = _build_consolidation_nudge("terminal_run", 3)
    assert n4 is not None and "commands" in n4
    # 非目标工具永不提醒
    assert _build_consolidation_nudge("unknown_tool", 99) is None


async def check_code_exec_breaker():
    """v1.0.29：code_exec 硬熔断——超阈值进入「熔断-放行」交替，打断 REPL 式增量循环。"""
    from zhishu.core.agent.agent import (
        _code_exec_breaker_message, _code_exec_final_directive, _CODE_EXEC_HARD_LIMIT)

    # 未超阈值：消息/指令均为空字符串（调用方据此判定不熔断）
    assert _code_exec_breaker_message(_CODE_EXEC_HARD_LIMIT) == ""
    assert _code_exec_final_directive(_CODE_EXEC_HARD_LIMIT) == ""

    # 超阈值熔断消息：含次数、引导整合为单脚本 + from_file 桥接
    msg = _code_exec_breaker_message(_CODE_EXEC_HARD_LIMIT + 5)
    assert "熔断" in msg and "code_exec" in msg and "整合" in msg and "from_file" in msg \
        and "generate_excel" in msg

    # 超阈值放行指令：含「最后一次机会」与完整脚本指引
    d = _code_exec_final_directive(_CODE_EXEC_HARD_LIMIT + 7)
    assert "最后一次" in d and "完整" in d and "generate_excel" in d

    # 模拟交替状态机：前 8 次自由（不熔断），第 9 次放行、第 10 次熔断、第 11 次放行…
    count = 0
    granted = True
    refuse_turns, final_turns, free_turns = [], [], []
    for i in range(1, 21):
        count += 1
        if count <= _CODE_EXEC_HARD_LIMIT:
            free_turns.append(i)
        elif granted:
            granted = False
            final_turns.append(i)
        else:
            granted = True
            refuse_turns.append(i)
    assert free_turns == list(range(1, _CODE_EXEC_HARD_LIMIT + 1)), free_turns
    assert refuse_turns == [10, 12, 14, 16, 18, 20], refuse_turns
    assert final_turns == [9, 11, 13, 15, 17, 19], final_turns


async def check_read_file_breaker():
    """v1.0.30：read_file 硬熔断——单文件读取超阈值进入「熔断-放行」交替，paths 批量读取始终放行。"""
    from zhishu.core.agent.agent import (
        _read_file_breaker_message, _read_file_final_directive,
        _READ_FILE_HARD_LIMIT, _read_file_is_batch)

    # _read_file_is_batch 真值表：仅 paths 列表（>=2 个）算批量
    assert _read_file_is_batch({"paths": ["a.txt", "b.csv"]}) is True
    assert _read_file_is_batch({"paths": ["a.txt"]}) is False
    assert _read_file_is_batch({"path": "a.txt"}) is False
    assert _read_file_is_batch({}) is False
    assert _read_file_is_batch({"paths": "a.txt"}) is False  # 非列表

    # 未超阈值：消息/指令均为空字符串
    assert _read_file_breaker_message(_READ_FILE_HARD_LIMIT) == ""
    assert _read_file_final_directive(_READ_FILE_HARD_LIMIT) == ""

    # 超阈值熔断消息：含次数、引导改用 paths 批量
    msg = _read_file_breaker_message(_READ_FILE_HARD_LIMIT + 5)
    assert "熔断" in msg and "paths" in msg and "批量" in msg and "read_file" in msg

    # 超阈值放行指令：含「最后一次单读」与 paths 批量指引
    d = _read_file_final_directive(_READ_FILE_HARD_LIMIT + 7)
    assert "最后一次" in d and "paths" in d

    # 模拟运行期状态机（与 agent.py 内逻辑一致）：单文件读取前 LIMIT 次自由，
    # 之后进入「放行(追加强指令)-熔断」交替；paths 批量读取永远不计入、直接放行。
    single = 0
    granted = True
    refuse_turns, final_turns, free_turns = [], [], []

    def _next(args):
        nonlocal single, granted
        if _read_file_is_batch(args):
            return "batch"  # 永远放行
        single += 1
        if single <= _READ_FILE_HARD_LIMIT:
            free_turns.append(single)
            return "free"
        if granted:
            granted = False
            final_turns.append(single)
            return "final"
        granted = True
        refuse_turns.append(single)
        return "refuse"

    # 20 次全单文件读取：free 1..6，之后 7/9/11..放行、8/10/12..熔断
    for _ in range(20):
        _next({"path": "f.txt"})
    assert free_turns == list(range(1, _READ_FILE_HARD_LIMIT + 1)), free_turns
    assert refuse_turns == [8, 10, 12, 14, 16, 18, 20], refuse_turns
    assert final_turns == [7, 9, 11, 13, 15, 17, 19], final_turns

    # paths 批量读取始终放行，且不会把单文件计数清零（单读计数在整轮内单调）
    outcomes = []
    for args in [{"path": "a"}] * 6 + [{"paths": ["x", "y"]}] * 3 + [{"path": "b"}] * 3:
        outcomes.append(_next(args))
    assert outcomes[6:9] == ["batch", "batch", "batch"]
    # 第 7 个单读（整体第 7 个 path 单读）触发 final，第 8 个触发 refuse，第 9 个 final
    assert outcomes[9] == "final" and outcomes[10] == "refuse" and outcomes[11] == "final"


async def main():
    await check_read_file()
    await check_terminal()
    await check_code_exec()
    await check_generate_excel()
    await check_generate_excel_from_file()
    await check_speed_hints()
    await check_schema_list_and_scalar()
    await check_consolidation_nudge()
    await check_code_exec_breaker()
    await check_read_file_breaker()
    print("ALL_TESTS_PASSED")


def test_batch_read_file():
    asyncio.run(check_read_file())


def test_batch_terminal():
    asyncio.run(check_terminal())


def test_batch_code_exec():
    asyncio.run(check_code_exec())


def test_batch_generate_excel():
    asyncio.run(check_generate_excel())


def test_generate_excel_from_file():
    asyncio.run(check_generate_excel_from_file())


def test_consolidation_nudge():
    asyncio.run(check_consolidation_nudge())


def test_code_exec_breaker():
    asyncio.run(check_code_exec_breaker())


def test_read_file_breaker():
    asyncio.run(check_read_file_breaker())


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except AssertionError as e:
        print("TEST_FAILED:", e)
        sys.exit(1)
    sys.exit(0)
