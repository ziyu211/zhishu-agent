"""生成真正的 Excel(.xlsx) 文件并返回 /media 下载链接。

重要：.xlsx 本质是 ZIP 包（含 [Content_Types].xml / xl/workbook.xml /
xl/worksheets/sheet1.xml / xl/_rels/workbook.xml.rels 等必需部件）。
**绝不能**用 file_write（文本模式）写 .xlsx，也不能手写 zip/XML 字节，
否则会生成 Excel 无法打开的「残缺 xlsx」。本工具由服务端用 openpyxl
在二进制模式正确生成合法 .xlsx，再落盘返回可下载链接。
"""
from __future__ import annotations

import os
import io
import csv
import json
import tempfile

from ..base import tool
from .sandbox import sandbox_cwd_for


def _coerce(v):
    """把单元格值规整为 openpyxl 可接受的标量（数字保持数字，其余转字符串）。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        return v
    return str(v)


def _build_sheet(ws, spec: dict):
    # CSV 文本输入优先（自动按逗号/制表符切分）
    csv_text = spec.get("csv")
    if csv_text and isinstance(csv_text, str):
        try:
            dialect = csv.Sniffer().sniff(csv_text[:4096] or ",", delimiters=",\t;")
        except Exception:
            dialect = csv.excel
        for row in csv.reader(io.StringIO(csv_text), dialect):
            ws.append([_coerce(c) for c in row])
        return
    header = spec.get("header")
    if header:
        ws.append([_coerce(h) for h in header])
    rows = spec.get("rows") or []
    for row in rows:
        if not isinstance(row, (list, tuple)):
            row = [row]
        ws.append([_coerce(c) for c in row])


# ── from_file 桥接（v1.0.28）：让模型在 code_exec 里算出并落盘的表格文件，
#    能直接转成合法 xlsx，解决「模型只 print 数据、generate_excel 拿不到数据」的断裂。 ──
def _resolve_from_file(from_file, media, owner):
    """把 from_file（绝对路径 / /media 链接 / 沙箱相对文件名）解析为可读的真实文件路径。"""
    if not from_file or not isinstance(from_file, str):
        return None
    f = from_file.strip()
    if not f:
        return None
    # 1) 绝对路径
    if os.path.isabs(f) and os.path.isfile(f):
        return f
    # 2) /media/... 链接 -> media.root 下的真实文件
    if f.startswith("/media/"):
        parts = f[len("/media/"):].strip("/").split("/")
        root = getattr(media, "root", None)
        if root:
            cand = os.path.join(root, *parts)
            if os.path.isfile(cand):
                return cand
    # 3) 沙箱相对文件名（模型在 code_exec 里写到 sandbox/<owner>/ 的文件）
    try:
        sb = sandbox_cwd_for(owner)
        cand = os.path.join(sb, os.path.basename(f))
        if os.path.isfile(cand):
            return cand
    except Exception:
        pass
    return None


def _load_sheet_spec_from_file(path: str) -> list | None:
    """从 CSV/JSON 文件构造 sheet spec 列表（[{name, header?, rows?, csv?}, ...]）。

    支持：
      * .csv/.tsv/.txt  -> 纯 CSV 文本（交给 _build_sheet 自动嗅探分隔符）
      * .json 且为 dict -> {表名: 行列表 | {header, rows}}
      * .json 且为 list -> list[list] 视为单表行；list[dict] 自动提取表头
    解析失败或为空返回 None。
    """
    if not path or not os.path.isfile(path):
        return None
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".json":
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                specs = []
                for name, body in data.items():
                    if isinstance(body, dict):
                        specs.append({"name": str(name)[:31],
                                      "header": body.get("header"),
                                      "rows": body.get("rows") or []})
                    elif isinstance(body, list):
                        specs.append({"name": str(name)[:31], "rows": body})
                return specs or None
            if isinstance(data, list):
                if data and isinstance(data[0], list):
                    return [{"name": "Sheet1", "rows": data}]
                if data and isinstance(data[0], dict):
                    header = list(data[0].keys())
                    rows = [[d.get(h, "") for h in header] for d in data]
                    return [{"name": "Sheet1", "header": header, "rows": rows}]
            return None
        # 默认按 CSV 处理
        with open(path, "r", encoding="utf-8", errors="ignore") as fh:
            csv_text = fh.read()
        if not csv_text.strip():
            return None
        return [{"name": os.path.splitext(os.path.basename(path))[0][:31], "csv": csv_text}]
    except Exception:
        return None


def _find_recent_sandbox_table(owner: str):
    """在用户沙箱目录里找最近修改的表格文件（csv/tsv/json/xlsx/xls），用于「只传 filename」的兜底。"""
    try:
        sb = sandbox_cwd_for(owner)
    except Exception:
        return None
    cands = []
    for fn in os.listdir(sb):
        if fn.lower().endswith((".csv", ".tsv", ".json", ".xlsx", ".xls")):
            fp = os.path.join(sb, fn)
            if os.path.isfile(fp):
                try:
                    cands.append((os.path.getmtime(fp), fp))
                except OSError:
                    continue
    if not cands:
        return None
    cands.sort(reverse=True)
    return cands[0][1]


@tool(
    "generate_excel",
    "【提速关键】需要一次生成『多个』Excel 文件时，必须用 files 列表（每项含 filename 与 sheet/sheets），"
    "一次调用返回多个 /media 下载链接（例：files:[{\"filename\":\"表1.xlsx\",\"sheet\":{...}},{\"filename\":\"表2.xlsx\",\"sheet\":{...}}]）。仅生成单文件才用 filename+sheet。"
    "生成真正合法的 Excel(.xlsx) 文件并返回 /media/... 可点击下载链接。"
    "当你需要『生成一张 Excel 表』（如销售表、库存表、导出数据、对比结果）时必须使用本工具，"
    "不要自己手写 xlsx 字节、不要拼 zip、也绝不要用 file_write 写 .xlsx（那会生成 Excel 打不开的损坏文件）。"
    "传入表格数据（表头+行，或多个工作表），服务端用 openpyxl 在二进制模式正确生成标准 .xlsx。",
    {
        "type": "object",
        "properties": {
            "filename": {"type": "string", "description": "单工作簿输出文件名，如 销售表.xlsx（缺省 export.xlsx）"},
            "sheet": {
                "type": "object",
                "description": "单个工作表：{name?, header?:[列名], rows?:[[...],...], csv?:'csv文本'}",
            },
            "sheets": {
                "type": "array",
                "description": "多个工作表：[{name?, header?, rows?, csv?}, ...]",
                "items": {"type": "object"},
            },
            "files": {
                "type": "array",
                "description": "批量生成多个独立 Excel 工作簿：[{filename?, sheet?/sheets?, from_file?}, ...]，一次调用返回多个 /media 下载链接",
                "items": {"type": "object"},
            },
            "from_file": {
                "type": "string",
                "description": "可选：表格数据来源文件（CSV/JSON/TXT）。支持 ① 绝对路径；② code_exec 返回的 /media/... 链接；③ 沙箱相对文件名（如 result.csv）。"
                "推荐做法：在 code_exec 里把表格写成 CSV 文件（如 open('result.csv','w',encoding='utf-8').write(csv_text)），它会自动发布并返回 /media 链接，再把该链接作为 from_file 传入。"
                "留空时自动采用本用户沙箱最近生成的表格文件。与 sheet/sheets/csv 互斥，传入则优先用该文件。",
            },
        },
        "required": [],
    },
    toolset="files",
)
async def generate_excel(args: dict, ctx) -> str:
    media = getattr(ctx, "media", None)
    if media is None:
        return "[generate_excel] 当前环境不支持（无媒体存储）"

    try:
        from openpyxl import Workbook
    except Exception:
        return "[generate_excel] 缺少 openpyxl 依赖，无法生成 Excel（请联系管理员安装）"

    owner = getattr(ctx, "user", "anonymous") or "anonymous"

    # 批量模式：files = [{filename?, sheet?/sheets?}, ...]
    files = args.get("files") or []
    if isinstance(files, str):
        files = [files]
    if not files:
        # 兼容单工作簿模式（filename + sheet/sheets / from_file）
        single_sheets = args.get("sheets")
        single_sheet = args.get("sheet")
        if not single_sheets and isinstance(single_sheet, dict):
            single_sheets = [single_sheet]
        files = [{
            "filename": (args.get("filename") or "export.xlsx").strip(),
            "sheets": single_sheets,
            "from_file": args.get("from_file"),
        }]

    def _save_one(spec: dict) -> str:
        if not isinstance(spec, dict):
            return f"[generate_excel] 无效的 files 元素: {spec!r}"
        filename = (spec.get("filename") or "export.xlsx").strip()
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            filename += ".xlsx"
        sheets_in = spec.get("sheets")
        single = spec.get("sheet")
        if not sheets_in and isinstance(single, dict):
            sheets_in = [single]

        # —— v1.0.28：from_file 桥接 + 沙箱最近表格文件兜底 ——
        auto_note = ""
        if not sheets_in or not isinstance(sheets_in, list):
            src = _resolve_from_file(spec.get("from_file"), media, owner)
            if src:
                loaded = _load_sheet_spec_from_file(src)
                if loaded:
                    sheets_in = loaded
            # 仅当调用方未显式给 from_file 时，才回退到沙箱最近表格文件（避免覆盖显式意图）
            if (not sheets_in or not isinstance(sheets_in, list)) and not spec.get("from_file"):
                sb_file = _find_recent_sandbox_table(owner)
                if sb_file:
                    loaded = _load_sheet_spec_from_file(sb_file)
                    if loaded:
                        sheets_in = loaded
                        auto_note = f"（自动采用沙箱最近生成的表格文件：{os.path.basename(sb_file)}）"
        if not sheets_in or not isinstance(sheets_in, list):
            return (
                f"[generate_excel] 「{filename}」缺少表格数据：请传入 sheets=[...] / "
                f"sheet={{...}} / csv 字段，或用 from_file='<CSV/JSON 文件绝对路径 或 /media 链接 或 沙箱文件名>'。"
                f"推荐：在 code_exec 中把表格写成 CSV 文件（如 open('result.csv','w',encoding='utf-8').write(csv_text)），"
                f"它会自动发布并返回 /media 链接，再把该链接作为 from_file 传入本工具即可生成合法 xlsx。"
            )
        wb = Workbook()
        # 移除默认空表，按用户数据重建
        default = wb.active
        wb.remove(default)
        for idx, s in enumerate(sheets_in):
            if not isinstance(s, dict):
                continue
            name = (s.get("name") or f"Sheet{idx + 1}")[:31]
            ws = wb.create_sheet(title=name)
            try:
                _build_sheet(ws, s)
            except Exception as e:  # noqa: BLE001
                ws.append([f"[该工作表解析失败: {e}]"])
        if not wb.sheetnames:
            wb.create_sheet("Sheet1")
        tmp_path = None
        try:
            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            wb.save(tmp_path)
            # 落盘到媒体库（保留文件名），返回 /media 链接
            url = media.save_file(tmp_path, kind="file", owner=owner)
        except Exception as e:  # noqa: BLE001
            return f"[generate_excel] 「{filename}」生成失败: {e}"
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
        n = len(wb.sheetnames)
        return (f"已生成标准 Excel 文件「{filename}」（共 {n} 个工作表，可被 Excel / WPS / libreoffice 正常打开）{auto_note}：\n"
                f"[{filename}]({url})")

    results = [_save_one(f) for f in files]
    if len(results) > 1:
        return "\n\n".join(results)
    return results[0] if results else "[generate_excel] 无有效生成任务"
